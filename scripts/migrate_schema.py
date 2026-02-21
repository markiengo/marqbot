"""
Cleanup utility for course workbook schema.

Removes deprecated `bucket1..bucket4` and `prereq_raw` columns from `courses`
after strict preflight checks and backup creation.

Usage:
    python scripts/migrate_schema.py --clean
    python scripts/migrate_schema.py --clean --path path/to/other.xlsx
    python scripts/migrate_schema.py --clean --dry-run
"""

import argparse
import os
import shutil
import sys

try:
    import openpyxl
    import pandas as pd
except ImportError as e:
    sys.exit(f"Missing dependency: {e}. Run: pip install openpyxl pandas")


DEFAULT_WORKBOOK = os.path.join(
    os.path.dirname(__file__), "..", "marquette_courses_full.xlsx"
)


def _sheet_headers(ws) -> list[str]:
    """Read first-row headers as stripped strings."""
    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    return [str(v).strip() if v is not None else "" for v in first_row]



def preflight_clean(wb: openpyxl.Workbook) -> None:
    """Abort unless cleanup preconditions are satisfied."""
    if "course_bucket" not in wb.sheetnames:
        sys.exit("[ERROR] 'course_bucket' sheet not found. Cleanup requires canonical mapping data.")
    if "courses" not in wb.sheetnames:
        sys.exit("[ERROR] No 'courses' sheet found in workbook.")

    map_ws = wb["course_bucket"]
    has_data_rows = any(
        any(val is not None and str(val).strip() for val in row)
        for row in map_ws.iter_rows(min_row=2, values_only=True)
    )
    if not has_data_rows:
        sys.exit("[ERROR] 'course_bucket' sheet has no data rows. Cleanup aborted.")


def find_deprecated_cols(courses_ws) -> list[int]:
    """
    Return 1-based openpyxl column indexes for deprecated columns.

    IMPORTANT: remove_columns() expects these indexes to be 1-based.
    """
    headers = _sheet_headers(courses_ws)
    deprecated = {"bucket1", "bucket2", "bucket3", "bucket4", "prereq_raw"}
    return sorted(
        idx for idx, name in enumerate(headers, start=1)
        if name in deprecated
    )


def backup_workbook(path: str) -> str:
    """Create/overwrite sibling .bak copy before destructive edits."""
    backup_path = f"{path}.bak"
    try:
        shutil.copy2(path, backup_path)
    except Exception as exc:
        sys.exit(f"[ERROR] Failed to create backup '{backup_path}': {exc}")
    print(f"[INFO] Backup created: {backup_path}")
    return backup_path


def remove_columns(courses_ws, col_indexes: list[int]) -> None:
    """Delete columns using 1-based openpyxl indexes, right-to-left."""
    for idx in sorted(col_indexes, reverse=True):
        courses_ws.delete_cols(idx, 1)


def main(args=None):
    parser = argparse.ArgumentParser(description="Clean deprecated columns from course workbook.")
    parser.add_argument("--path", default=DEFAULT_WORKBOOK, help="Path to the workbook")
    parser.add_argument("--dry-run", action="store_true", help="Preview without saving")
    parser.add_argument(
        "--clean",
        action="store_true",
        help="Remove deprecated columns from 'courses' after preflight + backup",
    )
    args = parser.parse_args(args)

    if not args.clean:
        print("[INFO] Only --clean mode is supported. Run with --clean flag.")
        return

    path = os.path.abspath(args.path)
    if not os.path.exists(path):
        sys.exit(f"[ERROR] Workbook not found: {path}")

    print(f"[INFO] Opening: {path}")
    wb = openpyxl.load_workbook(path)
    print(f"[INFO] Sheets: {wb.sheetnames}")

    preflight_clean(wb)
    courses_ws = wb["courses"]
    deprecated_col_indexes = find_deprecated_cols(courses_ws)

    if not deprecated_col_indexes:
        print("[INFO] No deprecated columns found. Nothing to remove.")
        return

    headers = _sheet_headers(courses_ws)
    deprecated_col_names = [
        headers[idx - 1] if idx - 1 < len(headers) else f"<col:{idx}>"
        for idx in deprecated_col_indexes
    ]

    if args.dry_run:
        print(f"[DRY RUN] Would remove deprecated columns: {deprecated_col_names}")
        print("[DRY RUN] No file changes were made.")
        return

    backup_path = backup_workbook(path)
    remove_columns(courses_ws, deprecated_col_indexes)
    wb.save(path)
    print(f"[DONE] Removed deprecated columns: {deprecated_col_names}")
    print(f"[DONE] Backup written: {backup_path}")
    print(f"[DONE] Saved: {path}")


if __name__ == "__main__":
    main()
