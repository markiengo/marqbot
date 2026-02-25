"""
Regression tests for schema normalization, legacy compatibility, and migration utilities.

Covers:
  1. _safe_bool_col() coercion behavior
  2. allocate_courses() policy-based double-count behavior
  3. data_loader legacy/new schema requirements
  4. migrate_schema --clean mode
  5. parent/child migration + elective purge
"""

import os

import openpyxl
import pandas as pd
import pytest

# conftest.py adds backend/ and scripts/ to sys.path.
from allocator import allocate_courses
from data_loader import _safe_bool_col, load_data
from migrate_schema import main as migrate_main
from migrate_parent_child_model import main as migrate_parent_child_main


# -----------------------------------------------------------------------------
# 1) Boolean coercion
# -----------------------------------------------------------------------------

class TestSafeBoolCol:
    def _apply(self, values):
        df = pd.DataFrame({"col": values})
        df = _safe_bool_col(df, "col")
        return df["col"].tolist()

    def test_python_bool_true(self):
        assert self._apply([True]) == [True]

    def test_python_bool_false(self):
        assert self._apply([False]) == [False]

    def test_int_1_is_true(self):
        assert self._apply([1]) == [True]

    def test_int_0_is_false(self):
        assert self._apply([0]) == [False]

    def test_float_1_is_true(self):
        assert self._apply([1.0]) == [True]

    def test_float_0_is_false(self):
        assert self._apply([0.0]) == [False]

    def test_string_TRUE(self):
        assert self._apply(["TRUE"]) == [True]

    def test_string_true_lowercase(self):
        assert self._apply(["true"]) == [True]

    def test_string_FALSE(self):
        assert self._apply(["FALSE"]) == [False]

    def test_string_false_lowercase(self):
        assert self._apply(["false"]) == [False]

    def test_string_1(self):
        assert self._apply(["1"]) == [True]

    def test_string_0(self):
        assert self._apply(["0"]) == [False]

    def test_string_yes(self):
        assert self._apply(["yes"]) == [True]

    def test_string_no(self):
        assert self._apply(["no"]) == [False]

    def test_string_y(self):
        assert self._apply(["y"]) == [True]

    def test_string_n(self):
        assert self._apply(["n"]) == [False]

    def test_nan_is_false(self):
        assert self._apply([float("nan")]) == [False]

    def test_none_is_false(self):
        assert self._apply([None]) == [False]

    def test_mixed_column(self):
        result = self._apply([1, 0, "TRUE", "false", True, None])
        assert result == [True, False, True, False, True, False]

    def test_missing_column_is_noop(self):
        df = pd.DataFrame({"other": [1, 2]})
        df2 = _safe_bool_col(df, "nonexistent")
        assert list(df2.columns) == ["other"]


# -----------------------------------------------------------------------------
# 2) Policy-driven double-count behavior
# -----------------------------------------------------------------------------

@pytest.fixture
def buckets_dc():
    return pd.DataFrame([
        {
            "track_id": "FIN_MAJOR",
            "bucket_id": "FIN_CHOOSE_2",
            "bucket_label": "Choose Two",
            "priority": 2,
            "needed_count": 2,
            "needed_credits": None,
            "min_level": 3000,
            "parent_bucket_id": "FIN_REQ",
        },
        {
            "track_id": "FIN_MAJOR",
            "bucket_id": "FIN_CHOOSE_1",
            "bucket_label": "Choose One",
            "priority": 3,
            "needed_count": 1,
            "needed_credits": None,
            "min_level": 3000,
            "parent_bucket_id": "FIN_REQ",
        },
    ])


@pytest.fixture
def map_no_can_double_count():
    return pd.DataFrame([
        {"track_id": "FIN_MAJOR", "bucket_id": "FIN_CHOOSE_2", "course_code": "FINA 4020"},
        {"track_id": "FIN_MAJOR", "bucket_id": "FIN_CHOOSE_1", "course_code": "FINA 4020"},
    ])


@pytest.fixture
def courses_dc():
    return pd.DataFrame([
        {"course_code": "FINA 4020", "course_name": "Financial Planning", "credits": 3, "level": 4000},
    ])


class TestPolicyDoubleCount:
    def test_double_count_works_with_policy_pair(
        self, buckets_dc, map_no_can_double_count, courses_dc
    ):
        policy = pd.DataFrame([
            {
                "program_id": "FIN_MAJOR",
                "node_type_a": "sub_bucket",
                "node_id_a": "FIN_CHOOSE_1",
                "node_type_b": "sub_bucket",
                "node_id_b": "FIN_CHOOSE_2",
                "allow_double_count": True,
            }
        ])
        result = allocate_courses(
            ["FINA 4020"],
            [],
            buckets_dc,
            map_no_can_double_count,
            courses_dc,
            double_count_policy_df=policy,
        )
        dc = result["double_counted_courses"]
        assert any(d["course_code"] == "FINA 4020" for d in dc)
        entry = next(d for d in dc if d["course_code"] == "FINA 4020")
        assert "FIN_CHOOSE_2" in entry["buckets"]
        assert "FIN_CHOOSE_1" in entry["buckets"]

    def test_double_count_blocked_without_policy_pair(self, buckets_dc, map_no_can_double_count, courses_dc):
        result = allocate_courses(
            ["FINA 4020"],
            [],
            buckets_dc,
            map_no_can_double_count,
            courses_dc,
            double_count_policy_df=pd.DataFrame(),
        )
        assert result["double_counted_courses"] == []


# -----------------------------------------------------------------------------
# 3) Strict V2 data loader
# -----------------------------------------------------------------------------

def _base_v2_sheets():
    courses = pd.DataFrame([{
        "course_code": "FINA 3001",
        "course_name": "Intro Finance",
        "credits": 3,
        "offered_fall": 1,
        "offered_spring": 1,
        "offered_summer": 0,
        "prereq_hard": "none",
        "prereq_level": 0,
    }])
    programs = pd.DataFrame([{
        "program_id": "FIN_MAJOR",
        "program_label": "Finance Major",
        "kind": "major",
        "parent_major_id": "",
        "active": 1,
    }, {
        "program_id": "CB",
        "program_label": "Corporate Banking",
        "kind": "track",
        "parent_major_id": "FIN_MAJOR",
        "active": 1,
    }])
    buckets = pd.DataFrame([{
        "program_id": "FIN_MAJOR",
        "bucket_id": "FIN_REQ",
        "bucket_label": "Finance Requirements",
        "priority": 1,
        "track_required": "",
        "active": 1,
    }])
    sub_buckets = pd.DataFrame([{
        "program_id": "FIN_MAJOR",
        "bucket_id": "FIN_REQ",
        "sub_bucket_id": "CORE",
        "sub_bucket_label": "Core",
        "courses_required": 1,
        "credits_required": None,
        "min_level": None,
        "role": "core",
        "priority": 1,
    }])
    course_sub_buckets = pd.DataFrame([{
        "program_id": "FIN_MAJOR",
        "sub_bucket_id": "CORE",
        "course_code": "FINA 3001",
        "constraints": "",
        "notes": "",
    }])
    double_count_policy = pd.DataFrame(columns=[
        "program_id",
        "node_type_a",
        "node_id_a",
        "node_type_b",
        "node_id_b",
        "allow_double_count",
    ])
    return (
        courses,
        programs,
        buckets,
        sub_buckets,
        course_sub_buckets,
        double_count_policy,
    )


class TestLoaderStrictMode:
    def test_loads_when_required_v2_sheets_present(self, tmp_path):
        (
            courses,
            programs,
            buckets,
            sub_buckets,
            course_sub_buckets,
            double_count_policy,
        ) = _base_v2_sheets()
        path = str(tmp_path / "test.xlsx")
        with pd.ExcelWriter(path, engine="openpyxl") as w:
            courses.to_excel(w, sheet_name="courses", index=False)
            programs.to_excel(w, sheet_name="programs", index=False)
            buckets.to_excel(w, sheet_name="buckets", index=False)
            sub_buckets.to_excel(w, sheet_name="sub_buckets", index=False)
            course_sub_buckets.to_excel(w, sheet_name="course_sub_buckets", index=False)
            double_count_policy.to_excel(w, sheet_name="double_count_policy", index=False)

        data = load_data(path)
        assert "CORE" in data["course_bucket_map_df"]["bucket_id"].values

    def test_raises_when_required_v2_sheet_missing(self, tmp_path):
        (
            courses,
            programs,
            buckets,
            sub_buckets,
            _course_sub_buckets,
            double_count_policy,
        ) = _base_v2_sheets()
        path = str(tmp_path / "test.xlsx")
        with pd.ExcelWriter(path, engine="openpyxl") as w:
            courses.to_excel(w, sheet_name="courses", index=False)
            programs.to_excel(w, sheet_name="programs", index=False)
            buckets.to_excel(w, sheet_name="buckets", index=False)
            sub_buckets.to_excel(w, sheet_name="sub_buckets", index=False)
            double_count_policy.to_excel(w, sheet_name="double_count_policy", index=False)

        with pytest.raises(ValueError, match="required map sheet"):
            load_data(path)

    def test_bool_coercion_on_loaded_data(self, tmp_path):
        (
            courses,
            programs,
            buckets,
            sub_buckets,
            course_sub_buckets,
            double_count_policy,
        ) = _base_v2_sheets()
        path = str(tmp_path / "test.xlsx")
        with pd.ExcelWriter(path, engine="openpyxl") as w:
            courses.to_excel(w, sheet_name="courses", index=False)
            programs.to_excel(w, sheet_name="programs", index=False)
            buckets.to_excel(w, sheet_name="buckets", index=False)
            sub_buckets.to_excel(w, sheet_name="sub_buckets", index=False)
            course_sub_buckets.to_excel(w, sheet_name="course_sub_buckets", index=False)
            double_count_policy.to_excel(w, sheet_name="double_count_policy", index=False)

        data = load_data(path)
        row = data["courses_df"].iloc[0]
        assert row["offered_fall"] == True
        assert row["offered_summer"] == False
        assert type(row["offered_fall"]) in (bool, __import__("numpy").bool_)

    def test_runtime_uses_priority_when_present(self, tmp_path):
        (
            courses,
            programs,
            buckets,
            sub_buckets,
            course_sub_buckets,
            double_count_policy,
        ) = _base_v2_sheets()
        sub_buckets.loc[:, "priority"] = 7
        path = str(tmp_path / "test.xlsx")
        with pd.ExcelWriter(path, engine="openpyxl") as w:
            courses.to_excel(w, sheet_name="courses", index=False)
            programs.to_excel(w, sheet_name="programs", index=False)
            buckets.to_excel(w, sheet_name="buckets", index=False)
            sub_buckets.to_excel(w, sheet_name="sub_buckets", index=False)
            course_sub_buckets.to_excel(w, sheet_name="course_sub_buckets", index=False)
            double_count_policy.to_excel(w, sheet_name="double_count_policy", index=False)

        data = load_data(path)
        row = data["buckets_df"][data["buckets_df"]["bucket_id"] == "CORE"].iloc[0]
        assert int(row["priority"]) == 7

    def test_runtime_derives_priority_when_sub_bucket_priority_missing(self, tmp_path):
        courses = pd.DataFrame([{
            "course_code": "FINA 3001",
            "course_name": "Intro Finance",
            "credits": 3,
            "offered_fall": 1,
            "offered_spring": 1,
            "offered_summer": 0,
            "prereq_hard": "none",
            "prereq_level": 0,
        }])
        programs = pd.DataFrame([{
            "program_id": "FIN_MAJOR",
            "program_label": "Finance Major",
            "kind": "major",
            "active": 1,
        }])
        buckets = pd.DataFrame([{
            "program_id": "FIN_MAJOR",
            "bucket_id": "FIN_REQ",
            "bucket_label": "Finance Requirements",
            "priority": 2,
            "track_required": "",
            "active": 1,
        }])
        sub_buckets = pd.DataFrame([
            {
                "program_id": "FIN_MAJOR",
                "bucket_id": "FIN_REQ",
                "sub_bucket_id": "CORE_A",
                "sub_bucket_label": "Core A",
                "courses_required": 1,
                "credits_required": None,
                "min_level": None,
                "role": "core",
            },
            {
                "program_id": "FIN_MAJOR",
                "bucket_id": "FIN_REQ",
                "sub_bucket_id": "ELEC_A",
                "sub_bucket_label": "Elec A",
                "courses_required": 1,
                "credits_required": None,
                "min_level": 3000,
                "role": "elective",
            },
        ])
        course_sub_buckets = pd.DataFrame([
            {"program_id": "FIN_MAJOR", "sub_bucket_id": "CORE_A", "course_code": "FINA 3001"},
            {"program_id": "FIN_MAJOR", "sub_bucket_id": "ELEC_A", "course_code": "FINA 3001"},
        ])
        double_count_policy = pd.DataFrame(columns=[
            "program_id",
            "node_type_a",
            "node_id_a",
            "node_type_b",
            "node_id_b",
            "allow_double_count",
        ])
        path = str(tmp_path / "test.xlsx")
        with pd.ExcelWriter(path, engine="openpyxl") as w:
            courses.to_excel(w, sheet_name="courses", index=False)
            programs.to_excel(w, sheet_name="programs", index=False)
            buckets.to_excel(w, sheet_name="buckets", index=False)
            sub_buckets.to_excel(w, sheet_name="sub_buckets", index=False)
            course_sub_buckets.to_excel(w, sheet_name="course_sub_buckets", index=False)
            double_count_policy.to_excel(w, sheet_name="double_count_policy", index=False)

        data = load_data(path)
        runtime = data["buckets_df"].set_index("bucket_id")
        core_p = int(runtime.loc["CORE_A", "priority"])
        elec_p = int(runtime.loc["ELEC_A", "priority"])
        assert core_p >= 200
        assert elec_p >= 200
        assert core_p < elec_p


# -----------------------------------------------------------------------------
# 4) --clean mode
# -----------------------------------------------------------------------------

def _make_clean_workbook(
    tmp_path,
    *,
    course_headers=("course_code", "bucket1", "bucket2", "extra_col"),
    course_row=("FINA 3001", "CORE", "ELECTIVE", "something"),
    include_course_sub_buckets=True,
    course_sub_buckets_has_data=True,
    filename="test.xlsx",
):
    wb = openpyxl.Workbook()
    del wb[wb.sheetnames[0]]

    ws_c = wb.create_sheet("courses")
    ws_c.append(list(course_headers))
    ws_c.append(list(course_row))

    if include_course_sub_buckets:
        ws_map = wb.create_sheet("course_sub_buckets")
        ws_map.append(["program_id", "sub_bucket_id", "course_code"])
        if course_sub_buckets_has_data:
            ws_map.append(["FIN_MAJOR", "CORE", "FINA 3001"])

    path = str(tmp_path / filename)
    wb.save(path)
    return path


def _courses_headers(path):
    with open(path, "rb") as f:
        wb = openpyxl.load_workbook(f)
    ws = wb["courses"]
    return [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]


class TestCleanMode:
    def test_clean_aborts_when_course_sub_buckets_missing(self, tmp_path):
        path = _make_clean_workbook(tmp_path, include_course_sub_buckets=False)
        headers_before = _courses_headers(path)
        with pytest.raises(SystemExit):
            migrate_main(["--clean", "--path", path])
        assert _courses_headers(path) == headers_before

    def test_clean_aborts_when_course_sub_buckets_empty(self, tmp_path):
        path = _make_clean_workbook(tmp_path, course_sub_buckets_has_data=False)
        headers_before = _courses_headers(path)
        with pytest.raises(SystemExit):
            migrate_main(["--clean", "--path", path])
        assert _courses_headers(path) == headers_before

    def test_clean_creates_backup_before_delete(self, tmp_path):
        path = _make_clean_workbook(tmp_path)
        migrate_main(["--clean", "--path", path])
        backup = path + ".bak"
        assert os.path.exists(backup)
        assert "bucket1" in _courses_headers(backup)
        assert "bucket1" not in _courses_headers(path)
        assert "bucket2" not in _courses_headers(path)

    def test_clean_noop_when_no_bucket_columns(self, tmp_path, capsys):
        path = _make_clean_workbook(
            tmp_path,
            course_headers=("course_code", "extra_col"),
            course_row=("FINA 3001", "something"),
        )
        headers_before = _courses_headers(path)
        migrate_main(["--clean", "--path", path])
        assert _courses_headers(path) == headers_before
        assert not os.path.exists(path + ".bak")
        assert "[INFO] No deprecated columns found. Nothing to remove." in capsys.readouterr().out

    def test_clean_dry_run_no_writes(self, tmp_path):
        path = _make_clean_workbook(tmp_path)
        migrate_main(["--clean", "--dry-run", "--path", path])
        assert not os.path.exists(path + ".bak")
        assert "bucket1" in _courses_headers(path)

    def test_clean_preserves_nondeprecated_columns_order(self, tmp_path):
        path = _make_clean_workbook(
            tmp_path,
            course_headers=("course_code", "bucket1", "extra_col", "bucket2"),
            course_row=("FINA 3001", "CORE", "something", "ELECTIVE"),
        )
        migrate_main(["--clean", "--path", path])
        headers = _courses_headers(path)
        assert "bucket1" not in headers
        assert "bucket2" not in headers
        assert headers == ["course_code", "extra_col"]


# -----------------------------------------------------------------------------
# 5) Parent/child migration + purge
# -----------------------------------------------------------------------------

def _legacy_parent_child_source():
    courses = pd.DataFrame([{
        "course_code": "FINA 3001",
        "course_name": "Intro Finance",
        "credits": 3,
        "offered_fall": 1,
        "offered_spring": 1,
        "offered_summer": 0,
        "prereq_hard": "none",
        "prereq_level": 0,
    }])
    programs = pd.DataFrame([
        {
            "program_id": "FIN_MAJOR",
            "program_label": "Finance Major",
            "kind": "major",
            "parent_major_id": "",
            "active": 1,
            "requires_primary_major": 0,
            "applies_to_all": 0,
        },
        {
            "program_id": "CB_TRACK",
            "program_label": "Commercial Banking",
            "kind": "track",
            "parent_major_id": "FIN_MAJOR",
            "active": 1,
            "requires_primary_major": 0,
            "applies_to_all": 0,
        },
        {
            "program_id": "BCC_CORE",
            "program_label": "Business Core Curriculum",
            "kind": "major",
            "parent_major_id": "",
            "active": 1,
            "requires_primary_major": 0,
            "applies_to_all": 1,
        },
    ])
    buckets = pd.DataFrame([
        {
            "program_id": "FIN_MAJOR",
            "bucket_id": "FIN_REQ",
            "bucket_label": "Finance Requirements",
            "priority": 1,
            "track_required": "",
            "active": 1,
        },
        {
            "program_id": "CB_TRACK",
            "bucket_id": "CB_REQ",
            "bucket_label": "Commercial Banking Requirements",
            "priority": 2,
            "track_required": "CB_TRACK",
            "active": 1,
        },
        {
            "program_id": "BCC_CORE",
            "bucket_id": "BCC",
            "bucket_label": "Business Core",
            "priority": 0,
            "track_required": "",
            "active": 1,
        },
    ])
    sub_buckets = pd.DataFrame([
        {
            "program_id": "FIN_MAJOR",
            "bucket_id": "FIN_REQ",
            "sub_bucket_id": "FIN_CORE",
            "sub_bucket_label": "Finance Core",
            "courses_required": 3,
            "credits_required": None,
            "min_level": None,
            "role": "core",
            "priority": 1,
        },
        {
            "program_id": "FIN_MAJOR",
            "bucket_id": "FIN_REQ",
            "sub_bucket_id": "FIN_CHOOSE_1",
            "sub_bucket_label": "Choose One",
            "courses_required": 1,
            "credits_required": None,
            "min_level": 3000,
            "role": "elective",
            "priority": 2,
        },
        {
            "program_id": "FIN_MAJOR",
            "bucket_id": "FIN_REQ",
            "sub_bucket_id": "FIN_BUS_ELEC_4",
            "sub_bucket_label": "Business Electives",
            "courses_required": None,
            "credits_required": 12,
            "min_level": 3000,
            "role": "elective",
            "priority": 3,
        },
        {
            "program_id": "FIN_MAJOR",
            "bucket_id": "FIN_REQ",
            "sub_bucket_id": "FIN_ELECTIVE_POOL",
            "sub_bucket_label": "Elective Pool",
            "courses_required": 2,
            "credits_required": None,
            "min_level": 3000,
            "role": "elective",
            "priority": 4,
        },
        {
            "program_id": "CB_TRACK",
            "bucket_id": "CB_REQ",
            "sub_bucket_id": "CB_CORE",
            "sub_bucket_label": "CB Core",
            "courses_required": 1,
            "credits_required": None,
            "min_level": None,
            "role": "core",
            "priority": 1,
        },
        {
            "program_id": "BCC_CORE",
            "bucket_id": "BCC",
            "sub_bucket_id": "BCC_FOUNDATION",
            "sub_bucket_label": "BCC Foundation",
            "courses_required": 1,
            "credits_required": None,
            "min_level": None,
            "role": "core",
            "priority": 1,
        },
    ])
    courses_all_buckets = pd.DataFrame([
        {"program_id": "FIN_MAJOR", "sub_bucket_id": "FIN_CORE", "course_code": "FINA 3001", "notes": ""},
        {"program_id": "FIN_MAJOR", "sub_bucket_id": "FIN_CHOOSE_1", "course_code": "FINA 4020", "notes": ""},
        {"program_id": "FIN_MAJOR", "sub_bucket_id": "fin_bus_elec_4", "course_code": "ACCO 1030", "notes": ""},
        {"program_id": "FIN_MAJOR", "sub_bucket_id": "fin_elective_pool", "course_code": "BUAD 1001", "notes": ""},
        {"program_id": "CB_TRACK", "sub_bucket_id": "CB_CORE", "course_code": "REAL 3001", "notes": ""},
        {"program_id": "BCC_CORE", "sub_bucket_id": "BCC_FOUNDATION", "course_code": "ECON 1001", "notes": ""},
    ])
    return courses, programs, buckets, sub_buckets, courses_all_buckets


class TestParentChildMigration:
    def test_migration_creates_parent_child_sheets_and_purges_elective_mappings(self, tmp_path):
        (
            courses,
            programs,
            buckets,
            sub_buckets,
            courses_all_buckets,
        ) = _legacy_parent_child_source()
        path = str(tmp_path / "parent_child_migration.xlsx")
        with pd.ExcelWriter(path, engine="openpyxl") as w:
            courses.to_excel(w, sheet_name="courses", index=False)
            programs.to_excel(w, sheet_name="programs", index=False)
            buckets.to_excel(w, sheet_name="buckets", index=False)
            sub_buckets.to_excel(w, sheet_name="sub_buckets", index=False)
            courses_all_buckets.to_excel(w, sheet_name="courses_all_buckets", index=False)

        migrate_parent_child_main(["--path", path])

        assert os.path.exists(path + ".bak")
        xl = pd.ExcelFile(path)
        assert "parent_buckets" in xl.sheet_names
        assert "child_buckets" in xl.sheet_names
        assert "master_bucket_courses" in xl.sheet_names

        parent = xl.parse("parent_buckets")
        child = xl.parse("child_buckets")
        master = xl.parse("master_bucket_courses")

        parent = parent.set_index("parent_bucket_id")
        assert str(parent.loc["BCC_CORE", "type"]).strip().lower() == "universal"
        assert bool(parent.loc["BCC_CORE", "active"]) is True
        assert str(parent.loc["CB_TRACK", "parent_major"]).strip().upper() == "FIN_MAJOR"

        child = child.set_index("child_bucket_id")
        assert str(child.loc["FIN_CORE", "requirement_mode"]).strip().lower() == "required"
        assert str(child.loc["FIN_CHOOSE_1", "requirement_mode"]).strip().lower() == "choose_n"
        assert str(child.loc["FIN_BUS_ELEC_4", "requirement_mode"]).strip().lower() == "credits_pool"

        mapped_child_ids = set(
            master["child_bucket_id"].fillna("").astype(str).str.strip().str.upper().tolist()
        )
        assert "FIN_CORE" in mapped_child_ids
        assert "FIN_CHOOSE_1" in mapped_child_ids
        assert "CB_CORE" in mapped_child_ids
        assert "FIN_BUS_ELEC_4" not in mapped_child_ids
        assert "FIN_ELECTIVE_POOL" not in mapped_child_ids

    def test_migration_dry_run_writes_nothing(self, tmp_path):
        (
            courses,
            programs,
            buckets,
            sub_buckets,
            courses_all_buckets,
        ) = _legacy_parent_child_source()
        path = str(tmp_path / "parent_child_dry_run.xlsx")
        with pd.ExcelWriter(path, engine="openpyxl") as w:
            courses.to_excel(w, sheet_name="courses", index=False)
            programs.to_excel(w, sheet_name="programs", index=False)
            buckets.to_excel(w, sheet_name="buckets", index=False)
            sub_buckets.to_excel(w, sheet_name="sub_buckets", index=False)
            courses_all_buckets.to_excel(w, sheet_name="courses_all_buckets", index=False)

        migrate_parent_child_main(["--dry-run", "--path", path])

        assert not os.path.exists(path + ".bak")
        xl = pd.ExcelFile(path)
        assert "parent_buckets" not in xl.sheet_names
        assert "child_buckets" not in xl.sheet_names
        assert "master_bucket_courses" not in xl.sheet_names


class TestParentChildLoaderCompatibility:
    def test_loader_accepts_parent_child_schema_and_normalizes_policy_aliases(self, tmp_path):
        courses = pd.DataFrame([{
            "course_code": "FINA 3001",
            "course_name": "Intro Finance",
            "credits": 3,
            "offered_fall": 1,
            "offered_spring": 1,
            "offered_summer": 0,
            "prereq_hard": "none",
            "prereq_level": 0,
        }])
        parent_buckets = pd.DataFrame([
            {
                "parent_bucket_id": "BCC_CORE",
                "parent_bucket_label": "Business Core",
                "type": "universal",
                "parent_major": "",
                "active": 1,
                "requires_primary_major": 0,
            },
            {
                "parent_bucket_id": "FIN_MAJOR",
                "parent_bucket_label": "Finance Major",
                "type": "major",
                "parent_major": "",
                "active": 1,
                "requires_primary_major": 0,
            },
            {
                "parent_bucket_id": "CB_TRACK",
                "parent_bucket_label": "Commercial Banking",
                "type": "track",
                "parent_major": "FIN_MAJOR",
                "active": 1,
                "requires_primary_major": 0,
            },
        ])
        child_buckets = pd.DataFrame([
            {
                "parent_bucket_id": "BCC_CORE",
                "child_bucket_id": "BCC_REQUIRED",
                "child_bucket_label": "BCC Required",
                "requirement_mode": "required",
                "courses_required": 1,
                "credits_required": None,
                "min_level": None,
                "notes": "",
            },
            {
                "parent_bucket_id": "FIN_MAJOR",
                "child_bucket_id": "FIN_CORE",
                "child_bucket_label": "Finance Core",
                "requirement_mode": "required",
                "courses_required": 1,
                "credits_required": None,
                "min_level": None,
                "notes": "",
            },
            {
                "parent_bucket_id": "FIN_MAJOR",
                "child_bucket_id": "FIN_BUS_ELEC_4",
                "child_bucket_label": "Business Electives",
                "requirement_mode": "credits_pool",
                "courses_required": None,
                "credits_required": 12,
                "min_level": 3000,
                "notes": "",
            },
            {
                "parent_bucket_id": "CB_TRACK",
                "child_bucket_id": "CB_CORE",
                "child_bucket_label": "CB Core",
                "requirement_mode": "required",
                "courses_required": 1,
                "credits_required": None,
                "min_level": None,
                "notes": "",
            },
        ])
        master_bucket_courses = pd.DataFrame([
            {
                "parent_bucket_id": "BCC_CORE",
                "child_bucket_id": "BCC_REQUIRED",
                "course_code": "ECON 1001",
                "notes": "",
            },
            {
                "parent_bucket_id": "FIN_MAJOR",
                "child_bucket_id": "FIN_CORE",
                "course_code": "FINA 3001",
                "notes": "",
            },
            {
                "parent_bucket_id": "FIN_MAJOR",
                "child_bucket_id": "FIN_BUS_ELEC_4",
                "course_code": "ACCO 1030",
                "notes": "",
            },
            {
                "parent_bucket_id": "CB_TRACK",
                "child_bucket_id": "CB_CORE",
                "course_code": "REAL 3001",
                "notes": "",
            },
        ])
        double_count_policy = pd.DataFrame([
            {
                "program_id": "FIN_MAJOR",
                "child_bucket_id_a": "FIN_CORE",
                "child_bucket_id_b": "FIN_BUS_ELEC_4",
                "allow_double_count": False,
                "reason": "same family",
            }
        ])
        path = str(tmp_path / "parent_child_loader.xlsx")
        with pd.ExcelWriter(path, engine="openpyxl") as w:
            courses.to_excel(w, sheet_name="courses", index=False)
            parent_buckets.to_excel(w, sheet_name="parent_buckets", index=False)
            child_buckets.to_excel(w, sheet_name="child_buckets", index=False)
            master_bucket_courses.to_excel(w, sheet_name="master_bucket_courses", index=False)
            double_count_policy.to_excel(w, sheet_name="double_count_policy", index=False)

        data = load_data(path)
        assert data["parent_child_detected"] is True

        v2_programs = data["v2_programs_df"].set_index("program_id")
        assert bool(v2_programs.loc["BCC_CORE", "applies_to_all"]) is True
        assert str(v2_programs.loc["CB_TRACK", "parent_major_id"]).strip().upper() == "FIN_MAJOR"

        mapped_child_ids = set(
            data["v2_courses_all_buckets_df"]["sub_bucket_id"]
            .fillna("")
            .astype(str)
            .str.strip()
            .str.upper()
            .tolist()
        )
        # Elective-like child mappings are purged at load time.
        assert "FIN_BUS_ELEC_4" not in mapped_child_ids
        assert "FIN_CORE" in mapped_child_ids

        policy_cols = set(data["v2_double_count_policy_df"].columns)
        assert "sub_bucket_id_a" in policy_cols
        assert "sub_bucket_id_b" in policy_cols


def test_loader_synthesizes_dynamic_elective_pool_mappings_from_tagged_courses(tmp_path):
    courses = pd.DataFrame([
        {
            "course_code": "FINA 3001",
            "course_name": "Intro Finance",
            "credits": 3,
            "level": 3000,
            "offered_fall": 1,
            "offered_spring": 1,
            "offered_summer": 0,
            "prereq_hard": "none",
            "prereq_level": 0,
            "elective_pool_tag": "",
        },
        {
            "course_code": "ACCO 3001",
            "course_name": "Intermediate Accounting",
            "credits": 3,
            "level": 3000,
            "offered_fall": 1,
            "offered_spring": 1,
            "offered_summer": 0,
            "prereq_hard": "none",
            "prereq_level": 0,
            "elective_pool_tag": "biz_elective",
        },
        {
            "course_code": "BUAD 1001",
            "course_name": "Business Foundations",
            "credits": 3,
            "level": 1000,
            "offered_fall": 1,
            "offered_spring": 1,
            "offered_summer": 0,
            "prereq_hard": "none",
            "prereq_level": 0,
            "elective_pool_tag": "biz_elective",
        },
    ])
    parent_buckets = pd.DataFrame([
        {
            "parent_bucket_id": "FIN_MAJOR",
            "parent_bucket_label": "Finance Major",
            "type": "major",
            "parent_major": "",
            "active": 1,
            "requires_primary_major": 0,
        },
    ])
    child_buckets = pd.DataFrame([
        {
            "parent_bucket_id": "FIN_MAJOR",
            "child_bucket_id": "FIN_CORE",
            "child_bucket_label": "Finance Core",
            "requirement_mode": "required",
            "courses_required": 1,
            "credits_required": None,
            "min_level": None,
            "notes": "",
        },
        {
            "parent_bucket_id": "FIN_MAJOR",
            "child_bucket_id": "FIN_BUS_ELEC_4",
            "child_bucket_label": "Business Elective Pool",
            "requirement_mode": "credits_pool",
            "courses_required": None,
            "credits_required": 12,
            "min_level": 3000,
            "notes": "",
        },
    ])
    master_bucket_courses = pd.DataFrame([
        {
            "parent_bucket_id": "FIN_MAJOR",
            "child_bucket_id": "FIN_CORE",
            "course_code": "FINA 3001",
            "notes": "",
        },
    ])

    path = str(tmp_path / "dynamic_elective_pool.xlsx")
    with pd.ExcelWriter(path, engine="openpyxl") as w:
        courses.to_excel(w, sheet_name="courses", index=False)
        parent_buckets.to_excel(w, sheet_name="parent_buckets", index=False)
        child_buckets.to_excel(w, sheet_name="child_buckets", index=False)
        master_bucket_courses.to_excel(w, sheet_name="master_bucket_courses", index=False)

    data = load_data(path)
    runtime_map = data["course_bucket_map_df"].copy()
    runtime_map["bucket_id"] = runtime_map["bucket_id"].fillna("").astype(str).str.strip().str.upper()
    runtime_map["course_code"] = runtime_map["course_code"].fillna("").astype(str).str.strip().str.upper()

    # Existing explicit mapping remains.
    core_codes = set(
        runtime_map.loc[runtime_map["bucket_id"] == "FIN_CORE", "course_code"].tolist()
    )
    assert "FINA 3001" in core_codes

    # Dynamic pool mappings are synthesized from tagged courses and respect min_level.
    elec_codes = set(
        runtime_map.loc[runtime_map["bucket_id"] == "FIN_BUS_ELEC_4", "course_code"].tolist()
    )
    assert "ACCO 3001" in elec_codes
    assert "BUAD 1001" not in elec_codes


def test_workbook_bcc_required_excludes_equivalency_rows():
    repo_root = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
    workbook_path = os.path.join(repo_root, "marquette_courses_full.xlsx")
    if not os.path.exists(workbook_path):
        pytest.skip("Workbook not present in test environment.")

    xl = pd.ExcelFile(workbook_path)
    if "master_bucket_courses" not in xl.sheet_names:
        pytest.skip("master_bucket_courses sheet not present.")

    master = xl.parse("master_bucket_courses")
    parent = master.get("parent_bucket_id", pd.Series(dtype=str)).fillna("").astype(str).str.strip().str.upper()
    child = master.get("child_bucket_id", pd.Series(dtype=str)).fillna("").astype(str).str.strip().str.upper()
    course = master.get("course_code", pd.Series(dtype=str)).fillna("").astype(str).str.strip().str.upper()
    banned = {"ACCO 4050", "COMM 1100"}

    mask = (parent == "BCC_CORE") & (child == "BCC_REQUIRED") & course.isin(banned)
    assert int(mask.sum()) == 0
